import asyncio

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from starlette.requests import Request

from backend import seed_autumn_workshops_2026
from backend.app.main import app
from backend.app.models.availability_subscriber import AvailabilitySubscriber
from backend.app.models.booking import Booking
from backend.app.models.workshop import Workshop
from backend.app.routes import paypal, public
from backend.app.services import availability_alert_service, email_service
from backend.app.services.excel_service import generate_xlsx


WORKSHOPS = (
    (
        "canfaito-2026",
        "CANFAITO & CONERO 2026",
        "/Canfaito_Conero_2026/",
        "2026-11-07",
    ),
    (
        "foreste-casentinesi-2026",
        "FORESTE CASENTINESI 2026",
        "/Foreste_Casentinesi_2026/",
        "2026-11-28",
    ),
)


def _memory_session(*tables):
    engine = create_engine("sqlite:///:memory:")
    for table in tables:
        table.__table__.create(engine)
    return sessionmaker(bind=engine)


def test_autumn_seed_creates_eight_seat_workshops(monkeypatch):
    Session = _memory_session(Workshop)
    monkeypatch.setattr(seed_autumn_workshops_2026, "SessionLocal", Session)
    monkeypatch.setattr(seed_autumn_workshops_2026, "init_db", lambda: None)

    result = seed_autumn_workshops_2026.seed()

    assert result == {
        "canfaito-2026": "created",
        "foreste-casentinesi-2026": "created",
    }
    db = Session()
    workshops = {item.workshop_key: item for item in db.query(Workshop).all()}
    for workshop_key, _, details_url, start_date in WORKSHOPS:
        workshop = workshops[workshop_key]
        assert workshop.total_seats == 8
        assert workshop.available_seats == 8
        assert workshop.price_cents == 35000
        assert workshop.start_date == start_date
        assert workshop.details_url == details_url
        assert workshop.status == "active"
    db.close()


@pytest.mark.parametrize("workshop_key, marker, _, __", WORKSHOPS)
def test_info_email_has_distinct_workshop_marker(monkeypatch, workshop_key, marker, _, __):
    captured = {}

    def fake_send_email(recipient, subject, body):
        captured.update(recipient=recipient, subject=subject, body=body)
        return True, "ok"

    monkeypatch.setattr(public, "send_email", fake_send_email)
    with TestClient(app) as client:
        response = client.post(
            "/api/send-info-email",
            json={
                "name": "Mario Rossi",
                "email": "mario@example.com",
                "phone": "+39 333 0000000",
                "source": workshop_key,
                "subject": "Richiesta informazioni",
                "message": "Vorrei alcune informazioni.",
            },
        )

    assert response.status_code == 200
    assert f"[{marker}]" in captured["subject"]
    assert f"Provenienza: [{marker}]" in captured["body"]
    assert "Email: mario@example.com" in captured["body"]


@pytest.mark.parametrize("workshop_key, _, __, ___", WORKSHOPS)
def test_payment_uses_same_deposit_and_total_without_friuli_extra(monkeypatch, workshop_key, _, __, ___):
    Session = _memory_session(Workshop, Booking)
    db = Session()
    db.add(
        Workshop(
            workshop_key=workshop_key,
            slug=workshop_key,
            title=f"Workshop {workshop_key}",
            total_seats=8,
            available_seats=8,
            price_cents=35000,
            status="active",
        )
    )
    db.commit()
    db.close()

    monkeypatch.setattr(paypal, "SessionLocal", Session)
    paypal_payload = {}

    def fake_paypal_request(method, path, body=None):
        paypal_payload.update(body or {})
        return {
            "id": f"ORDER-{workshop_key}",
            "links": [{"rel": "approve", "href": "https://paypal.example/approve"}],
        }

    monkeypatch.setattr(paypal, "_paypal_request", fake_paypal_request)
    body = paypal.CreateOrderRequest(
        workshopId=workshop_key,
        formula="saldo",
        extraDay=True,
        firstName="Mario",
        lastName="Rossi",
        email="mario@example.com",
        participants=1,
    )
    request = Request(
        {
            "type": "http",
            "method": "POST",
            "path": "/api/create-paypal-order",
            "headers": [],
            "client": ("127.0.0.1", 12345),
        }
    )

    result = asyncio.run(paypal.create_paypal_order(body, request, None))

    check = Session()
    booking = check.query(Booking).one()
    assert booking.final_cents == 35000
    assert booking.amount_due_cents == 35000
    assert booking.balance_cents == 30000
    assert booking.extra_day_selected is False
    assert result["extraDay"] is False
    assert paypal_payload["purchase_units"][0]["amount"]["value"] == "350.00"
    check.close()


@pytest.mark.parametrize("workshop_key, marker, details_url, _", WORKSHOPS)
def test_availability_alert_uses_workshop_page(monkeypatch, workshop_key, marker, details_url, _):
    Session = _memory_session(Workshop, AvailabilitySubscriber)
    db = Session()
    workshop = Workshop(
        workshop_key=workshop_key,
        slug=workshop_key,
        title=f"Workshop {workshop_key}",
        total_seats=8,
        available_seats=2,
        price_cents=35000,
        details_url=details_url,
    )
    subscriber = AvailabilitySubscriber(
        workshop_id=workshop_key,
        name="Mario",
        email="mario@example.com",
        consent_at="2026-08-19T00:00:00+00:00",
        consent_source=f"{workshop_key}-landing",
        unsubscribe_token=f"token-{workshop_key}",
        active=True,
    )
    db.add_all([workshop, subscriber])
    db.commit()

    captured = {}

    def fake_send_email(recipient, subject, body):
        captured.update(recipient=recipient, subject=subject, body=body)
        return True, "ok"

    monkeypatch.setattr(availability_alert_service, "send_email", fake_send_email)
    monkeypatch.setattr(availability_alert_service.settings, "aruba_smtp_pass", "test-only")

    assert availability_alert_service.notify_availability_threshold(db, workshop) == 1
    assert f"[{workshop_key.replace('-', ' ').upper()}]" in captured["subject"]
    assert f"https://www.davideluongo.it{details_url}" in captured["body"]
    db.close()


def test_non_friuli_email_and_report_do_not_show_friday_option(monkeypatch, tmp_path):
    sent = []

    def fake_send_email(recipient, subject, body, attachment_path=None):
        sent.append(body)
        return True, "ok"

    monkeypatch.setattr(email_service, "send_email", fake_send_email)
    email_service.send_booking_confirmation(
        {
            "id": "BK-TESTAUTUMN",
            "firstName": "Mario",
            "lastName": "Rossi",
            "email": "mario@example.com",
            "formula": "caparra",
            "workshopId": "canfaito-2026",
            "workshopName": "Workshop Canfaito & Conero",
            "finalCents": 35000,
            "balanceCents": 30000,
            "extraDay": False,
        }
    )
    assert len(sent) == 2
    assert all("venerdì" not in body.lower() for body in sent)

    workshop = Workshop(
        workshop_key="canfaito-2026",
        slug="canfaito-2026",
        title="Workshop Canfaito & Conero",
        total_seats=8,
        available_seats=7,
        price_cents=35000,
        status="active",
    )
    booking = Booking(
        id="BK-TESTAUTUMN",
        status="paid",
        workshop_id="canfaito-2026",
        workshop_name=workshop.title,
        first_name="Mario",
        last_name="Rossi",
        email="mario@example.com",
        formula="caparra",
        final_cents=35000,
        balance_cents=30000,
        amount_due_cents=5000,
    )
    path, _ = generate_xlsx(workshop, [booking], output_dir=tmp_path)
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Partecipanti"]
    headers = [cell.value for cell in sheet[1]]
    friday_column = headers.index("Dal venerdì") + 1
    assert sheet.cell(2, friday_column).value == "—"
