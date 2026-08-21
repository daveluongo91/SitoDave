import asyncio

from starlette.requests import Request
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from backend.app.models.booking import Booking
from backend.app.models.workshop import Workshop
from backend.app.routes import paypal
from backend.app.services import email_service
from backend.app.services.excel_service import generate_xlsx
from openpyxl import load_workbook


def _create_order(monkeypatch, *, extra_day: bool, formula: str):
    engine = create_engine("sqlite:///:memory:")
    Workshop.__table__.create(engine)
    Booking.__table__.create(engine)
    Session = sessionmaker(bind=engine)
    db = Session()
    db.add(
        Workshop(
            workshop_key="friuli-2026",
            slug="friuli-2026",
            title="Workshop Friuli: Laghi e Cascate",
            total_seats=8,
            available_seats=8,
            price_cents=35000,
            status="active",
        )
    )
    db.commit()
    db.close()

    monkeypatch.setattr(paypal, "SessionLocal", Session)
    paypal_payload = {}

    def fake_paypal_request(method, path, body=None):
        paypal_payload.update(body or {})
        return {
            "id": f"ORDER-{formula}-{int(extra_day)}",
            "links": [{"rel": "approve", "href": "https://paypal.example/approve"}],
        }

    monkeypatch.setattr(paypal, "_paypal_request", fake_paypal_request)
    body = paypal.CreateOrderRequest(
        workshopId="friuli-2026",
        formula=formula,
        extraDay=extra_day,
        firstName="Mario",
        lastName="Rossi",
        email="mario@example.com",
        phone="+39 333 0000000",
        participants=1,
    )
    request = Request(
        {
            "type": "http",
            "method": "POST",
            "path": "/api/create-paypal-order",
            "headers": [],
            "client": ("127.0.0.1", 12345),
        }
    )
    result = asyncio.run(paypal.create_paypal_order(body, request, None))
    check = Session()
    booking = check.query(Booking).filter(Booking.paypal_order_id == result["orderId"]).one()
    values = {
        "result": result,
        "original": booking.original_cents,
        "final": booking.final_cents,
        "balance": booking.balance_cents,
        "due": booking.amount_due_cents,
        "extra": booking.extra_day_selected,
        "extra_cents": booking.extra_day_cents,
        "paypal_amount": paypal_payload["purchase_units"][0]["amount"]["value"],
    }
    check.close()
    return values


def test_friuli_standard_total_is_350(monkeypatch):
    values = _create_order(monkeypatch, extra_day=False, formula="saldo")
    assert values["original"] == 35000
    assert values["due"] == 35000
    assert values["balance"] == 30000
    assert values["extra"] is False
    assert values["result"]["finalPrice"] == "350.00"
    assert values["paypal_amount"] == "350.00"


def test_friuli_extra_day_total_is_450(monkeypatch):
    values = _create_order(monkeypatch, extra_day=True, formula="saldo")
    assert values["original"] == 45000
    assert values["final"] == 45000
    assert values["due"] == 45000
    assert values["balance"] == 40000
    assert values["extra"] is True
    assert values["extra_cents"] == 10000
    assert values["result"]["extraDayAmount"] == "100.00"
    assert values["paypal_amount"] == "450.00"


def test_friuli_extra_day_keeps_deposit_at_50(monkeypatch):
    values = _create_order(monkeypatch, extra_day=True, formula="caparra")
    assert values["final"] == 45000
    assert values["due"] == 5000
    assert values["balance"] == 40000
    assert values["paypal_amount"] == "50.00"


def test_confirmation_email_mentions_friday(monkeypatch):
    sent = []

    def fake_send_email(recipient, subject, body, attachment_path=None):
        sent.append((recipient, subject, body))
        return True, "ok"

    monkeypatch.setattr(email_service, "send_email", fake_send_email)
    email_service.send_booking_confirmation(
        {
            "id": "BK-TESTFRIDAY",
            "firstName": "Mario",
            "lastName": "Rossi",
            "email": "mario@example.com",
            "formula": "saldo",
            "workshopId": "friuli-2026",
            "workshopName": "Workshop Friuli: Laghi e Cascate",
            "finalCents": 45000,
            "balanceCents": 40000,
            "extraDay": True,
        }
    )

    assert len(sent) == 2
    assert all("venerdì 9 ottobre (+€100)" in body for _, _, body in sent)


def test_cutoff_report_contains_friday_option(tmp_path):
    workshop = Workshop(
        workshop_key="friuli-2026",
        slug="friuli-2026",
        title="Workshop Friuli: Laghi e Cascate",
        start_date="2026-10-10",
        end_date="2026-10-11",
        total_seats=8,
        available_seats=7,
        price_cents=35000,
        status="active",
    )
    booking = Booking(
        id="BK-TESTEXTRA01",
        status="paid",
        workshop_id="friuli-2026",
        workshop_name=workshop.title,
        first_name="Mario",
        last_name="Rossi",
        email="mario@example.com",
        phone="+39 333 0000000",
        formula="saldo",
        extra_day_selected=True,
        extra_day_cents=10000,
        original_cents=45000,
        final_cents=45000,
        balance_cents=40000,
        amount_due_cents=45000,
    )

    path, _ = generate_xlsx(workshop, [booking], output_dir=tmp_path)
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Partecipanti"]
    headers = [cell.value for cell in sheet[1]]
    friday_column = headers.index("Dal venerdì") + 1
    assert sheet.cell(2, friday_column).value == "Sì (+€100)"
    assert sheet.cell(2, headers.index("Importo Previsto (€)") + 1).value == 450
