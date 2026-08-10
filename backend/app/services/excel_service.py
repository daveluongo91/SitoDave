"""
backend/app/services/excel_service.py
Generazione XLSX reale con openpyxl.
Anti formula injection, formati italiani, intestazioni, filtri, riga bloccata.
"""
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from backend.app.config.settings import settings
from backend.app.models.booking import Booking
from backend.app.models.workshop import Workshop
from backend.app.models.cost import WorkshopCost


# ── Anti formula injection ────────────────────────────────────────────────────

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value: str) -> str:
    """
    Previene formula injection nei valori stringa.
    Antepone un apostrofo ai valori che iniziano con caratteri pericolosi.
    """
    if not isinstance(value, str):
        return value
    v = value.strip()
    if v and v[0] in _FORMULA_PREFIXES:
        return "'" + v
    return v


def _euro(value: str | Decimal | None) -> Decimal:
    """Converte a Decimal sicuro."""
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


# ── Stili ─────────────────────────────────────────────────────────────────────

def _header_style():
    """Stile intestazione: sfondo blu scuro, testo bianco, grassetto."""
    fill = PatternFill("solid", fgColor="0B1929")
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return fill, font, align


def _apply_header(ws, row: int, headers: list[str]) -> None:
    fill, font, align = _header_style()
    thin = Side(style="thin", color="2A5A8A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Generazione XLSX ──────────────────────────────────────────────────────────

def generate_xlsx(
    workshop: Workshop,
    bookings: list[Booking],
    cost: Optional[WorkshopCost] = None,
    output_dir: Optional[Path] = None,
) -> tuple[Path, str]:
    """
    Genera un file XLSX reale con tre fogli: Partecipanti, Riepilogo, Costi.
    Restituisce (filepath, sha256_hash).
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl non installato. Esegui: pip install openpyxl")

    out_dir = output_dir or settings.exports_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"report_{workshop.workshop_key}_{ts}.xlsx"
    filepath = out_dir / filename

    wb = Workbook()

    # ── Foglio 1: Partecipanti ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Partecipanti"

    headers_p = [
        "ID Prenotazione", "Nome", "Cognome", "Email", "Telefono",
        "Workshop", "Data Iscrizione", "Modalità Pagamento",
        "Importo Previsto (€)", "Saldo Pagato", "Stato Amministrativo",
        "Codice Sconto", "Sconto (€)", "Note", "Privacy Accettata",
    ]
    _apply_header(ws1, 1, headers_p)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers_p))}1"

    fmt_date = "DD/MM/YYYY"
    fmt_euro = '#,##0.00 "€"'

    for row_idx, b in enumerate(bookings, start=2):
        alt = PatternFill("solid", fgColor="F0F4F8") if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        final_eur = _euro(b.final_cents) / Decimal("100") if b.final_cents else Decimal("0")
        disc_eur = _euro(b.discount_cents) / Decimal("100") if b.discount_cents else Decimal("0")

        row_data = [
            _safe_cell(b.id or ""),
            _safe_cell(b.first_name or ""),
            _safe_cell(b.last_name or ""),
            _safe_cell(b.email or ""),
            _safe_cell(b.phone or ""),
            _safe_cell(b.workshop_name or workshop.title),
            b.created_at[:10] if b.created_at else "",      # data ISO → stringa
            _safe_cell(b.formula or ""),
            float(final_eur),
            "Sì" if b.balance_paid else "No",
            _safe_cell(b.admin_status or b.status or ""),
            _safe_cell(b.coupon_code or ""),
            float(disc_eur),
            _safe_cell(b.admin_notes or ""),
            "Sì" if b.privacy_accepted else "No",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = alt
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Formato valuta per colonne importi
            if col_idx in (9, 13):
                cell.number_format = fmt_euro

    _set_col_widths(ws1, [14, 16, 16, 28, 16, 20, 13, 18, 16, 12, 18, 14, 12, 30, 14])

    # ── Foglio 2: Riepilogo ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Riepilogo")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 30

    paid_count = sum(1 for b in bookings if b.status == "paid" and not b.is_deleted)
    total_revenue = sum(
        (_euro(b.final_cents) / Decimal("100"))
        for b in bookings if b.status == "paid" and not b.is_deleted
    )
    remaining = max(0, workshop.total_seats - paid_count)

    cost_total = _euro(cost.total_costs_decimal) if cost else Decimal("0")
    margin = total_revenue - cost_total

    summary = [
        ("RIEPILOGO WORKSHOP", ""),
        ("Workshop", workshop.title),
        ("Data inizio", workshop.start_date or ""),
        ("Data fine", workshop.end_date or ""),
        ("Cutoff", workshop.cutoff_at or "Non impostato"),
        ("Posti totali", workshop.total_seats),
        ("Partecipanti (paganti)", paid_count),
        ("Posti rimanenti", remaining),
        ("Totale ricavi previsti (€)", float(total_revenue)),
        ("Totale costi (€)", float(cost_total)),
        ("Margine stimato (€)", float(margin)),
        ("Data report", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")),
    ]

    fill_header, font_header, _ = _header_style()
    for row_idx, (label, value) in enumerate(summary, start=1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = Font(bold=True, size=12, color="0B1929")
        elif isinstance(value, float):
            cell_b.number_format = '#,##0.00 "€"'

    # ── Foglio 3: Costi ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Costi")
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 20

    if cost:
        cost_items = [
            ("COSTI ORGANIZZATIVI", ""),
            ("Pernottamenti (notti)", cost.nights),
            ("Costo per notte (€)", float(_euro(cost.cost_per_night_decimal))),
            ("Camere/persone", cost.room_count),
            ("Totale pernottamento (€)", float(_euro(cost.total_accommodation_decimal))),
            ("", ""),
            ("Carburante (€)", float(_euro(cost.fuel_decimal))),
            ("Pedaggi (€)", float(_euro(cost.tolls_decimal))),
            ("Parcheggi (€)", float(_euro(cost.parking_decimal))),
            ("Traghetti (€)", float(_euro(cost.ferries_decimal))),
            ("Altre spese viaggio (€)", float(_euro(cost.other_travel_decimal))),
            ("Totale viaggio (€)", float(_euro(cost.total_travel_decimal))),
            ("", ""),
            ("Altre spese organizzative (€)", float(_euro(cost.other_org_decimal))),
            ("TOTALE COSTI (€)", float(_euro(cost.total_costs_decimal))),
            ("Costo per partecipante (€)", float(_euro(cost.cost_per_participant_decimal))),
            ("Margine stimato (€)", float(_euro(cost.estimated_margin_decimal))),
            ("", ""),
            ("Fonte stima", cost.estimate_source or "Inserimento manuale"),
            ("Data verifica", cost.verified_at or ""),
            ("Link ViaMichelin", cost.viamichelin_url or ""),
        ]
    else:
        cost_items = [("Nessun dato costi disponibile", "")]

    fmt_euro3 = '#,##0.00 "€"'
    for row_idx, (label, value) in enumerate(cost_items, start=1):
        cell_a = ws3.cell(row=row_idx, column=1, value=label)
        cell_b = ws3.cell(row=row_idx, column=2, value=value)
        if label.startswith("TOTALE") or label.startswith("COSTI"):
            cell_a.font = Font(bold=True)
        if isinstance(value, float):
            cell_b.number_format = fmt_euro3

    wb.save(filepath)

    # Calcola hash SHA256
    file_bytes = filepath.read_bytes()
    file_hash = hashlib.sha256(file_bytes).hexdigest()

    return filepath, file_hash
