"""
backend/app/services/export_service.py
Generazione pura e riutilizzabile di export partecipanti in formato XLSX.
Non modifica stato di cutoff, non tocca prenotazioni, applica filtri e protezione formula injection.
"""
from __future__ import annotations

import hashlib
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import List, Optional

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from backend.app.models.booking import Booking
from backend.app.models.workshop import Workshop
from backend.app.models.cost import WorkshopCost
from backend.app.services.excel_service import _safe_cell, _euro


def generate_participants_export_xlsx(
    workshop: Optional[Workshop],
    bookings: List[Booking],
    cost: Optional[WorkshopCost] = None,
    filter_label: str = "Tutti",
    output_dir: Optional[Path] = None,
) -> tuple[Path, str, str]:
    """
    Genera un file Excel di export estemporaneo per i partecipanti.
    Restituisce (filepath, sha256_hash, filename).
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl non e' installato.")

    now = datetime.now(timezone.utc)
    ws_key = workshop.workshop_key if workshop else "tutti_i_workshop"
    ws_title = workshop.title if workshop else "Tutti i Workshop & Esperienze"
    
    timestamp_str = now.strftime("%Y-%m-%d_%H%M")
    filename = f"partecipanti_{ws_key}_{timestamp_str}.xlsx"

    if output_dir is None:
        from backend.app.config.settings import settings
        output_dir = settings.exports_dir

    output_dir.mkdir(parents=True, exist_ok=True)
    filepath = output_dir / filename

    wb = Workbook()
    
    # ── 1. Foglio Partecipanti ────────────────────────────────────────────────
    ws_part = wb.active
    ws_part.title = "Partecipanti"
    ws_part.views.sheetView[0].showGridLines = True

    # Titolo del report
    title_font = Font(name="Calibri", size=14, bold=True, color="0B1929")
    sub_font = Font(name="Calibri", size=10, italic=True, color="555555")
    
    ws_part["A1"] = _safe_cell(f"LISTA PARTECIPANTI — {ws_title.upper()}")
    ws_part["A1"].font = title_font
    ws_part["A2"] = _safe_cell(f"Esportazione: {now.strftime('%d/%m/%Y %H:%M UTC')} • Filtro applicato: {filter_label}")
    ws_part["A2"].font = sub_font

    headers = [
        "ID Prenotazione",
        "Data Iscrizione",
        "Workshop / Esperienza",
        "Nome",
        "Cognome",
        "Email",
        "Telefono",
        "Pax",
        "Formula",
        "Giorno Extra",
        "Stato Pagamento",
        "Quota Listino (€)",
        "Sconto (€)",
        "Versato Online (€)",
        "Saldo Dovuto (€)",
        "Saldo Pagato in Loco",
        "Metodo Saldo",
        "Note Admin",
    ]

    header_row = 4
    header_fill = PatternFill("solid", fgColor="0B1929")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_part.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Blocca i riquadri sotto l'intestazione
    ws_part.freeze_panes = f"A{header_row + 1}"

    # Dati
    row_idx = header_row + 1
    total_participants = 0
    total_original = Decimal("0")
    total_discount = Decimal("0")
    total_paid_online = Decimal("0")
    total_balance_due = Decimal("0")

    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    num_fmt_euro = '#,##0.00 "€"'

    for b in bookings:
        total_participants += (b.participants or 1)
        orig_dec = Decimal(b.original_cents or 0) / 100
        disc_dec = Decimal(b.discount_cents or 0) / 100
        final_dec = Decimal(b.final_cents or 0) / 100
        bal_due_dec = Decimal(b.balance_cents or 0) / 100

        total_original += orig_dec
        total_discount += disc_dec
        total_paid_online += final_dec
        total_balance_due += bal_due_dec

        row_data = [
            _safe_cell(b.id),
            _safe_cell(b.created_at[:16].replace("T", " ") if b.created_at else ""),
            _safe_cell(b.workshop_name or b.workshop_id or ""),
            _safe_cell(b.first_name),
            _safe_cell(b.last_name),
            _safe_cell(b.email),
            _safe_cell(str(b.phone or "")),
            b.participants or 1,
            _safe_cell("Caparra" if b.formula == "caparra" else ("Saldo" if b.formula == "saldo" else (b.formula or ""))),
            "Sì" if b.extra_day_selected else "No",
            _safe_cell(b.status.upper()),
            orig_dec,
            disc_dec,
            final_dec,
            bal_due_dec,
            "Sì" if b.balance_paid else "No",
            _safe_cell(b.balance_paid_method or "-"),
            _safe_cell(b.admin_notes or ""),
        ]

        use_zebra = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_part.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill

            if col_idx in (12, 13, 14, 15):
                cell.number_format = num_fmt_euro
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (8,):
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (1, 2, 9, 10, 11, 16, 17):
                cell.alignment = Alignment(horizontal="center")

        row_idx += 1

    # Riga Totali
    if bookings:
        total_row = row_idx
        ws_part.cell(row=total_row, column=1, value="TOTALE COMPLESSIVO").font = Font(name="Calibri", size=10, bold=True)
        ws_part.cell(row=total_row, column=8, value=total_participants).font = Font(name="Calibri", size=10, bold=True)
        ws_part.cell(row=total_row, column=12, value=total_original).number_format = num_fmt_euro
        ws_part.cell(row=total_row, column=13, value=total_discount).number_format = num_fmt_euro
        ws_part.cell(row=total_row, column=14, value=total_paid_online).number_format = num_fmt_euro
        ws_part.cell(row=total_row, column=15, value=total_balance_due).number_format = num_fmt_euro

        for col in (12, 13, 14, 15):
            ws_part.cell(row=total_row, column=col).font = Font(name="Calibri", size=10, bold=True)
            ws_part.cell(row=total_row, column=col).alignment = Alignment(horizontal="right")

        for col in range(1, len(headers) + 1):
            cell = ws_part.cell(row=total_row, column=col)
            cell.border = Border(top=Side(style="double", color="0B1929"), bottom=Side(style="double", color="0B1929"))

    # Auto-fit larghezza colonne
    for col in ws_part.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < header_row:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_part.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Abilita filtri automatici su tutta l'intestazione
    last_col_letter = get_column_letter(len(headers))
    ws_part.auto_filter.ref = f"A{header_row}:{last_col_letter}{max(row_idx - 1, header_row)}"

    # ── 2. Foglio Riepilogo ───────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Riepilogo")
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary["A1"] = _safe_cell(f"RIEPILOGO STATISTICO — {ws_title.upper()}")
    ws_summary["A1"].font = title_font

    summary_items = [
        ("Numero Prenotazioni", len(bookings)),
        ("Totale Partecipanti Effettivi (Pax)", total_participants),
        ("Totale Valore Listino", total_original),
        ("Totale Sconti Erogati", total_discount),
        ("Totale Incassato Online", total_paid_online),
        ("Totale Saldo da Riscuotere", total_balance_due),
        ("Stima Incasso Complessivo", total_paid_online + total_balance_due),
    ]

    ws_summary.cell(row=3, column=1, value="Metrica").font = Font(bold=True)
    ws_summary.cell(row=3, column=2, value="Valore").font = Font(bold=True)

    for s_idx, (k, v) in enumerate(summary_items, start=4):
        ws_summary.cell(row=s_idx, column=1, value=k)
        c_val = ws_summary.cell(row=s_idx, column=2, value=v)
        if isinstance(v, Decimal):
            c_val.number_format = num_fmt_euro
            c_val.alignment = Alignment(horizontal="right")
        else:
            c_val.alignment = Alignment(horizontal="center")

    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 24

    # ── 3. Foglio Costi (se presente) ─────────────────────────────────────────
    if cost:
        ws_cost = wb.create_sheet(title="Costi Workshop")
        ws_cost.views.sheetView[0].showGridLines = True
        ws_cost["A1"] = _safe_cell("PIANO COSTI E MARGINI")
        ws_cost["A1"].font = title_font

        cost_rows = [
            ("Notti Soggiorno", cost.nights),
            ("Costo a Notte (€)", _euro(cost.cost_per_night_decimal)),
            ("Numero Camere", cost.room_count),
            ("Totale Alloggio (€)", _euro(cost.total_accommodation_decimal)),
            ("Carburante (€)", _euro(cost.fuel_decimal)),
            ("Pedaggi (€)", _euro(cost.tolls_decimal)),
            ("Parcheggi (€)", _euro(cost.parking_decimal)),
            ("Traghetti (€)", _euro(cost.ferries_decimal)),
            ("Altri Costi Viaggio (€)", _euro(cost.other_travel_decimal)),
            ("Altri Costi Organizzativi (€)", _euro(cost.other_org_decimal)),
            ("Totale Costi Stimati (€)", _euro(cost.total_costs_decimal)),
            ("Costo Stimato per Partecipante (€)", _euro(cost.cost_per_participant_decimal)),
            ("Margine Stimato (€)", _euro(cost.estimated_margin_decimal)),
        ]

        for c_idx, (ck, cv) in enumerate(cost_rows, start=4):
            ws_cost.cell(row=c_idx, column=1, value=ck)
            c_cell = ws_cost.cell(row=c_idx, column=2, value=cv)
            if isinstance(cv, Decimal):
                c_cell.number_format = num_fmt_euro
                c_cell.alignment = Alignment(horizontal="right")
            else:
                c_cell.alignment = Alignment(horizontal="center")

        ws_cost.column_dimensions["A"].width = 36
        ws_cost.column_dimensions["B"].width = 22

    wb.save(filepath)

    with open(filepath, "rb") as f:
        file_hash = hashlib.sha256(f.read()).hexdigest()

    return filepath, file_hash, filename